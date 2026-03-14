from flask import Flask, render_template, request, redirect, url_for, send_from_directory, session
import os, io, hashlib
from datetime import datetime
from PIL import Image
from backend.db import hospitals_collection, reports_collection
from backend.hospital_model import create_hospital, get_hospital_by_id, verify_hospital_login
from backend.hash_utils import generate_sha256
from backend.blockchain import store_hash, get_hash
from backend.qr_utils import generate_qr, get_report_id_from_token
from backend.report_metadata import store_metadata, get_metadata, get_reports_by_hospital
from backend.verification_logger import log_verification, get_verification_logs
from reportlab.pdfgen import canvas
from openpyxl import Workbook
from openpyxl.styles import Font
import pikepdf
import requests
from datetime import datetime
import pytz
import re
import pdfplumber
import difflib
# ================= GET REAL IP =================
def get_real_ip():

    if request.headers.get("X-Forwarded-For"):
        return request.headers.get("X-Forwarded-For").split(",")[0]

    if request.headers.get("X-Real-IP"):
        return request.headers.get("X-Real-IP")

    return request.remote_addr


# ================= GET LOCATION FROM IP =================
def get_location_from_ip(ip):

    try:

        res = requests.get(f"http://ip-api.com/json/{ip}", timeout=5)

        data = res.json()

        if data["status"] == "success":

            city = data.get("city", "")
            country = data.get("country", "")

            return f"{city}, {country}"

        return "Unknown"

    except:
        return "Unknown"
    
# ================= GET CITY, COUNTRY FROM LAT LONG =================
def get_location_from_coordinates(latitude, longitude, ip_address=None):

    try:
        if not latitude or not longitude:
            return get_location_from_ip(ip_address) if ip_address else "Unknown"

        url = "https://nominatim.openstreetmap.org/reverse"

        params = {
            "lat": latitude,
            "lon": longitude,
            "format": "json",
            "addressdetails": 1
        }

        headers = {
            "User-Agent": "MedLedger Medical Verification (contact@medledger.com)"
        }

        response = requests.get(
            url,
            params=params,
            headers=headers,
            timeout=8
        )

        if response.status_code == 200:

            data = response.json()
            address = data.get("address", {})

            city = (
                address.get("city")
                or address.get("town")
                or address.get("village")
                or address.get("municipality")
                or address.get("county")
                or address.get("state_district")
                or address.get("state")
            )

            country = address.get("country")

            if city and country:
                return f"{city}, {country}"

        # 🔁 FALLBACK TO IP LOCATION
        if ip_address:
            return get_location_from_ip(ip_address)

        return f"{latitude}, {longitude}"

    except Exception as e:
        print("GPS Location error:", e)

        if ip_address:
            return get_location_from_ip(ip_address)

        return f"{latitude}, {longitude}"
    
# ======================================================
# APP CONFIG
# ======================================================
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

app = Flask(
    __name__,
    template_folder=os.path.join(BASE_DIR, "templates"),
    static_folder=os.path.join(BASE_DIR, "static")
)

app.secret_key = "medchain_secret_key"

UPLOAD_FOLDER = os.path.join(os.getcwd(), "backend", "uploads")
QR_DIR = os.path.join(BASE_DIR, "static", "qrcodes")

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(QR_DIR, exist_ok=True)

ADMIN_USERNAME = "admin"
ADMIN_PASSWORD = "admin123"

ALLOWED_EXTENSIONS = (".pdf", ".png", ".jpg", ".jpeg")

# ======================================================
# APPROVAL REQUIRED DECORATOR
# ======================================================
from functools import wraps

def approval_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):

        if "hospital_id" not in session:
            return redirect("/login")

        hospital = get_hospital_by_id(session["hospital_id"])

        if not hospital:
            session.clear()
            return redirect("/login")

        # check approval
        if hospital["role"] == "main":
            if hospital.get("admin_status") != "approved":
                return render_template("pending_approval.html", hospital=hospital)

        elif hospital["role"] == "branch":
            if (
                hospital.get("admin_status") != "approved"
                or hospital.get("main_status") != "approved"
            ):
                return render_template("pending_approval.html", hospital=hospital)

        return f(*args, **kwargs)

    return decorated_function

# ======================================================
# APPROVAL CHECK FUNCTION
# ======================================================
def is_approved(hospital):

    if hospital["role"] == "main":
        return hospital.get("admin_status") == "approved"

    if hospital["role"] == "branch":
        return (
            hospital.get("admin_status") == "approved"
            and hospital.get("main_status") == "approved"
        )

    return False

# ======================================================
# PDF QR EMBED
# ======================================================
def attach_qr_to_pdf_safe(pdf_path, qr_path, output_path):
    pdf = pikepdf.open(pdf_path)
    last_page = pdf.pages[-1]

    w, h = float(last_page.MediaBox[2]), float(last_page.MediaBox[3])

    packet = io.BytesIO()
    c = canvas.Canvas(packet, pagesize=(w, h))
    size = min(w, h) * 0.15

    c.drawImage(qr_path, w - size - 20, 20, size, size)
    c.save()

    packet.seek(0)
    overlay_path = output_path.replace(".pdf", "_overlay.pdf")

    with open(overlay_path, "wb") as f:
        f.write(packet.read())

    with pikepdf.open(overlay_path) as overlay:
        last_page.add_overlay(overlay.pages[0])

    pdf.save(output_path)
    pdf.close()
    os.remove(overlay_path)


# ======================================================
# IMAGE QR EMBED
# ======================================================
def embed_qr_into_image(image_path, qr_path, output_path):
    image = Image.open(image_path).convert("RGB")
    qr = Image.open(qr_path).convert("RGB")

    img_w, img_h = image.size
    qr_size = int(min(img_w, img_h) * 0.25)
    qr = qr.resize((qr_size, qr_size))

    position = (
        img_w - qr_size - 20,
        img_h - qr_size - 20
    )

    image.paste(qr, position)
    image.save(output_path)


# ======================================================
# HOME
# ======================================================
@app.route("/")
def home():
    return render_template("home.html")


# ======================================================
# ADMIN AUTH
# ======================================================
@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    if request.method == "POST":
        if (
            request.form.get("username") == ADMIN_USERNAME and
            request.form.get("password") == ADMIN_PASSWORD
        ):
            session["admin"] = True
            session["admin_login_success"] = True
            return redirect("/admin/dashboard")

        return render_template("admin/admin_login.html",
                               error="Invalid admin credentials")

    return render_template("admin/admin_login.html")


@app.route("/admin/logout")
def admin_logout():

    # remove admin session
    session.clear()

    # set logout notification flag
    session["logout_success"] = True

    # redirect to home page
    return redirect(url_for("home"))


# ======================================================
# ADMIN DASHBOARD
# ======================================================
@app.route("/admin/dashboard")
def admin_dashboard():

    if not session.get("admin"):
        return redirect("/admin/login")

    hospitals = list(hospitals_collection.find({"role": "main"}))

    # ================= GET VERIFICATION STATS =================
    logs = get_verification_logs({})

    genuine_count = 0
    fake_count = 0

    for log in logs:

        if log.get("result") == "Genuine":
            genuine_count += 1

        elif log.get("result") == "Fake":
            fake_count += 1

    total_verifications = genuine_count + fake_count

    return render_template(
        "admin/admin_dashboard.html",
        hospitals=hospitals,
        genuine_count=genuine_count,
        fake_count=fake_count,
        total_verifications=total_verifications
    )

@app.route("/admin/main/<main_id>")
def admin_view_main(main_id):

    if not session.get("admin"):
        return redirect("/admin/login")

    main_hospital = hospitals_collection.find_one({
        "hospital_id": main_id,
        "role": "main"
    })

    branches = list(hospitals_collection.find({
        "parent_hospital": main_id,
        "role": "branch"
    }))

    reports = list(reports_collection.find({
        "hospital_id": {"$in": [b["hospital_id"] for b in branches]}
    }).sort("created_at", -1))

    return render_template(
        "admin/admin_main_details.html",
        main=main_hospital,
        branches=branches,
        reports=reports
    )

@app.route("/admin/approve/<hospital_id>")
def approve_hospital(hospital_id):

    if not session.get("admin"):
        return redirect("/admin/login")

    hospitals_collection.update_one(
        {"hospital_id": hospital_id},
        {"$set": {"admin_status": "approved"}}
    )

    return redirect("/admin/dashboard?success=approved")

@app.route("/admin/reject/<hospital_id>")
def reject_hospital(hospital_id):

    if not session.get("admin"):
        return redirect("/admin/login")

    hospitals_collection.update_one(
        {"hospital_id": hospital_id},
        {"$set": {"admin_status": "pending"}}
    )

    return redirect("/admin/dashboard?success=rejected")


@app.route("/admin/delete/<hospital_id>")
def delete_hospital(hospital_id):

    if not session.get("admin"):
        return redirect("/admin/login")

    hospitals_collection.delete_one({"hospital_id": hospital_id})
    reports_collection.delete_many({"hospital_id": hospital_id})

    return redirect("/admin/dashboard?success=deleted")

@app.route("/admin/branch/approve/<branch_id>")
def admin_approve_branch(branch_id):

    if not session.get("admin"):
        return redirect("/admin/login")

    hospitals_collection.update_one(
        {"hospital_id": branch_id},
        {"$set": {"admin_status": "approved"}}
    )

    return redirect(request.referrer.split("?")[0] + "?success=branch_approved")

@app.route("/admin/branch/reject/<branch_id>")
def admin_reject_branch(branch_id):

    if not session.get("admin"):
        return redirect("/admin/login")

    hospitals_collection.update_one(
        {"hospital_id": branch_id, "role": "branch"},
        {"$set": {"admin_status": "pending"}}
    )

    return redirect(request.referrer.split("?")[0] + "?success=branch_rejected")

@app.route("/admin/branch/delete/<branch_id>")
def admin_delete_branch(branch_id):

    if not session.get("admin"):
        return redirect("/admin/login")

    hospitals_collection.delete_one({
        "hospital_id": branch_id,
        "role": "branch"
    })

    reports_collection.delete_many({
        "hospital_id": branch_id
    })

    return redirect(request.referrer.split("?")[0] + "?success=branch_deleted")

@app.route("/admin/reports/<hospital_id>")
def admin_view_reports(hospital_id):

    if not session.get("admin"):
        return redirect("/admin/login")

    reports = list(reports_collection.find({
        "hospital_id": hospital_id
    }).sort("uploaded_at", -1))

    enriched_reports = []

    for report in reports:

        hospital = hospitals_collection.find_one({
            "hospital_id": report.get("hospital_id")
        })

        hospital_name_full = "Unknown"
        hospital_address = ""

        if hospital:

            hospital_name_full = hospital.get("hospital_name", "")
            hospital_address = hospital.get("address", "")

            if hospital.get("branch_name"):
                hospital_name_full += f" – {hospital.get('branch_name')}"
            else:
                hospital_name_full += " – Main"

        report["hospital_name_full"] = hospital_name_full
        report["hospital_address"] = hospital_address

        enriched_reports.append(report)

    return render_template(
        "admin/admin_reports.html",
        reports=enriched_reports,
        hospital_id=hospital_id
    )

@app.route("/download/admin-reports-excel/<hospital_id>")
def download_admin_reports_excel(hospital_id):

    if not session.get("admin"):
        return redirect("/admin/login")

    reports = list(reports_collection.find({
        "hospital_id": hospital_id
    }).sort("uploaded_at", -1))

    wb = Workbook()
    ws = wb.active
    ws.title = "Hospital Reports"

    headers = [

        "Report ID",

        "Uploaded Hospital",
        "Hospital Address",

        "Upload Time",

        "Location",
        "IP Address",

        "Blockchain Hash",
        "File Name"
    ]

    ws.append(headers)

    for col in range(1, len(headers)+1):
        ws.cell(row=1, column=col).font = Font(bold=True)


    for report in reports:

        hospital = hospitals_collection.find_one({
            "hospital_id": report.get("hospital_id")
        })

        hospital_name_full = "Unknown"
        hospital_address = ""

        if hospital:

            hospital_name_full = hospital.get("hospital_name", "")
            hospital_address = hospital.get("address", "")

            if hospital.get("branch_name"):
                hospital_name_full += f" – {hospital.get('branch_name')}"
            else:
                hospital_name_full += " – Main"


        ws.append([

            report.get("report_id"),

            hospital_name_full,
            hospital_address,

            report.get("uploaded_at_str"),

            report.get("location"),
            report.get("ip_address"),

            report.get("hash"),

            report.get("file_name")

        ])

    filename = f"admin_reports_{hospital_id}.xlsx"
    filepath = os.path.join(UPLOAD_FOLDER, filename)

    wb.save(filepath)

    return send_from_directory(
        UPLOAD_FOLDER,
        filename,
        as_attachment=True
    )

@app.route("/admin/verifications")
def admin_my_verifications():

    if not session.get("admin"):
        return redirect("/admin/login")

    logs = get_verification_logs({
        "verifiedBy": "Admin"
    })

    enriched_logs = []

    for log in logs:

        # ================= UPLOADED HOSPITAL =================
        report_hospital = hospitals_collection.find_one({
            "hospital_id": log.get("reportHospitalId")
        })

        report_name = "Unknown"
        report_address = ""

        if report_hospital:

            report_name = report_hospital.get("hospital_name", "")
            report_address = report_hospital.get("address", "")

            if report_hospital.get("branch_name"):
                report_name += f" – {report_hospital.get('branch_name')}"
            else:
                report_name += " – Main"

        # ================= VERIFIED BY =================
        verifier_name = "System Admin"
        verifier_address = "System Network"

        # ================= ADD FIELDS =================
        log["reportHospitalNameFull"] = report_name
        log["reportHospitalAddress"] = report_address

        log["verifierNameFull"] = verifier_name
        log["verifierAddress"] = verifier_address

        enriched_logs.append(log)

    return render_template(
        "admin/admin_my_verifications.html",
        logs=enriched_logs
    )

@app.route("/download/admin-my-verifications-excel")
def download_admin_my_verifications_excel():

    if not session.get("admin"):
        return redirect("/admin/login")

    logs = get_verification_logs({
        "verifiedBy": "Admin"
    })

    wb = Workbook()
    ws = wb.active
    ws.title = "Admin Verification Logs"

    headers = [

        "Report ID",

        "Uploaded Hospital",
        "Hospital Address",

        "Verified By",
        "Verifier Address",

        "Verified At",

        "Location",
        "IP Address",

        "Original Hash",
        "Scanned Hash",

        "Result",

        "Original File",
        "Scanned File"
    ]

    ws.append(headers)

    for col in range(1, len(headers)+1):
        ws.cell(row=1, column=col).font = Font(bold=True)

    for log in logs:

        report_hospital = hospitals_collection.find_one({
            "hospital_id": log.get("reportHospitalId")
        })

        report_name = ""
        report_address = ""

        if report_hospital:

            report_name = report_hospital.get("hospital_name","")
            report_address = report_hospital.get("address","")

            if report_hospital.get("branch_name"):
                report_name += f" – {report_hospital.get('branch_name')}"
            else:
                report_name += " – Main"


        ws.append([

            log.get("reportId"),

            report_name,
            report_address,

            "System Admin",
            "System",

            log["verifiedAt"].strftime("%d %b %Y %I:%M %p"),

            log.get("location"),
            log.get("ipAddress"),

            log.get("originalHash"),
            log.get("scannedHash"),

            log.get("result"),

            f"{log.get('reportId')}_final.pdf",
            log.get("scannedFile")

        ])

    filename = "admin_verification_logs.xlsx"

    path = os.path.join(UPLOAD_FOLDER, filename)

    wb.save(path)

    return send_from_directory(
        UPLOAD_FOLDER,
        filename,
        as_attachment=True
    )

@app.route("/admin/report/delete/<report_id>")
def admin_delete_report(report_id):

    if not session.get("admin"):
        return redirect("/admin/login")

    report = reports_collection.find_one({"report_id": report_id})

    if report:
        file_path = os.path.join(UPLOAD_FOLDER, report["file_name"])
        if os.path.exists(file_path):
            os.remove(file_path)

        reports_collection.delete_one({"_id": report["_id"]})

    return redirect(request.referrer.split("?")[0] + "?success=deleted")


# ======================================================
# HOSPITAL AUTH
# ======================================================
@app.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "POST":
        data = {
            "hospital_name": request.form.get("hospital_name"),
            "hospital_id": request.form.get("hospital_id"),
            "email": request.form.get("email"),
            "password": request.form.get("password"),
            "address": request.form.get("address"),
            "role": request.form.get("role"), # main / branch
            "branch_name": request.form.get("branch_name"), # only for branch
            "parent_hospital": request.form.get("parent_hospital"), # main hospital id
            "admin_status": "pending",
"main_status": "pending"
        }

        if get_hospital_by_id(data["hospital_id"]):
            return render_template("register.html",
                                   error="Hospital already exists")

        create_hospital(data)
        session["register_success"] = True
        return redirect("/login")

    return render_template("register.html")


@app.route("/login", methods=["GET", "POST"])
def login():

    if request.method == "POST":

        hospital = verify_hospital_login(
            request.form.get("hospital_id"),
            request.form.get("password")
        )

        if not hospital:
            return render_template(
                "login.html",
                error="Invalid credentials"
            )

        # SAVE SESSION
        session["hospital_id"] = hospital["hospital_id"]
        session["hospital_name"] = hospital["hospital_name"]
        session["role"] = hospital.get("role")
        session["branch_name"] = hospital.get("branch_name")
        session["parent_hospital"] = hospital.get("parent_hospital")

        # SAVE APPROVAL STATUS (VERY IMPORTANT)
        session["admin_status"] = hospital.get("admin_status")
        session["main_status"] = hospital.get("main_status")

        # REDIRECT ALWAYS TO DASHBOARD
        if hospital["role"] == "main":
            session["main_login_success"] = True
            return redirect("/main-dashboard")

        if hospital["role"] == "branch":
            session["login_success"] = True
            return redirect("/hospital-dashboard")

    return render_template("login.html")


@app.route("/logout")
def logout():

    # clear all session data
    session.clear()

    # set logout notification flag
    session["logout_success"] = True

    # redirect to home page
    return redirect(url_for("home"))



# ======================================================
# HOSPITAL DASHBOARD
# ======================================================
@app.route("/hospital-dashboard")
@approval_required
def hospital_dashboard():

    hospital = get_hospital_by_id(session["hospital_id"])

    reports = list(reports_collection.find({
        "hospital_id": session["hospital_id"]
    }).sort("uploaded_at", -1))

    return render_template(
        "hospital_dashboard.html",
        hospital=hospital,
        reports=reports
    )

@app.route("/hospital/report/delete/<report_id>")
def delete_report(report_id):

    if "hospital_id" not in session:
        return redirect("/login")

    # ❌ BLOCK DELETE
    return redirect("/hospital-dashboard?error=delete_not_allowed")

@app.route("/hospital/verifications")
@approval_required
def my_branch_verifications():

    if session.get("role") != "branch":
        return redirect("/login")

    logs = get_verification_logs({
        "verifiedBy": "Hospital",
        "verifiedByHospitalId": session["hospital_id"]
    })

    return render_template(
        "hospital_verifications.html",
        logs=logs
    )

@app.route("/hospital/my-verifications")
@approval_required
def my_verification_logs():

    hospital_id = session["hospital_id"]

    logs = get_verification_logs({
        "verifiedByHospitalId": hospital_id
    })

    enriched_logs = []

    for log in logs:

        # ================= UPLOADED HOSPITAL =================
        report_hospital = hospitals_collection.find_one({
            "hospital_id": log.get("reportHospitalId")
        })

        report_name = "Unknown"
        report_address = ""

        if report_hospital:

            report_name = report_hospital.get("hospital_name", "")
            report_address = report_hospital.get("address", "")

            if report_hospital.get("branch_name"):
                report_name += f" – {report_hospital.get('branch_name')}"
            else:
                report_name += " – Main"


        # ================= VERIFIED BY =================
        verifier_hospital = hospitals_collection.find_one({
            "hospital_id": log.get("verifiedByHospitalId")
        })

        verifier_name = log.get("verifiedByHospitalName", "Public User")
        verifier_address = ""

        if verifier_hospital:

            verifier_address = verifier_hospital.get("address", "")

            if verifier_hospital.get("branch_name"):
                verifier_name += f" – {verifier_hospital.get('branch_name')}"
            else:
                verifier_name += " – Main"

        else:

            verifier_name = "Public User"
            verifier_address = "Public Network"


        # ================= ADD FIELDS =================
        log["reportHospitalNameFull"] = report_name
        log["reportHospitalAddress"] = report_address

        log["verifierNameFull"] = verifier_name
        log["verifierAddress"] = verifier_address

        enriched_logs.append(log)


    return render_template(
        "hospital_my_verifications.html",
        logs=enriched_logs
    )

@app.route("/download/hospital-my-verifications-excel")
@approval_required
def download_hospital_my_verifications_excel():

    hospital_id = session["hospital_id"]

    logs = get_verification_logs({
        "verifiedByHospitalId": hospital_id
    })

    wb = Workbook()
    ws = wb.active
    ws.title = "My Verification Logs"

    headers = [

        "Report ID",

        "Uploaded Hospital",
        "Hospital Address",

        "Verified By",
        "Verifier Address",

        "Verified At",

        "Location",
        "IP Address",

        "Original Hash",
        "Scanned Hash",

        "Result",

        "Original File",
        "Scanned File"

    ]

    ws.append(headers)

    for col in range(1, len(headers)+1):
        ws.cell(row=1, column=col).font = Font(bold=True)


    for log in logs:

        # ================= UPLOADED HOSPITAL =================
        report_hospital = hospitals_collection.find_one({
            "hospital_id": log.get("reportHospitalId")
        })

        report_name = "Unknown"
        report_address = ""

        if report_hospital:

            report_name = report_hospital.get("hospital_name", "")
            report_address = report_hospital.get("address", "")

            if report_hospital.get("branch_name"):
                report_name += f" – {report_hospital.get('branch_name')}"
            else:
                report_name += " – Main"


        # ================= VERIFIED BY =================
        verifier_hospital = hospitals_collection.find_one({
            "hospital_id": log.get("verifiedByHospitalId")
        })

        verifier_name = log.get("verifiedByHospitalName", "Public User")
        verifier_address = ""

        if verifier_hospital:

            verifier_address = verifier_hospital.get("address", "")

            if verifier_hospital.get("branch_name"):
                verifier_name += f" – {verifier_hospital.get('branch_name')}"
            else:
                verifier_name += " – Main"

        else:

            verifier_name = "Public User"
            verifier_address = "Public Network"


        ws.append([

            log.get("reportId"),

            report_name,
            report_address,

            verifier_name,
            verifier_address,

            log["verifiedAt"].strftime("%d %b %Y %I:%M %p"),

            log.get("location"),
            log.get("ipAddress"),

            log.get("originalHash"),
            log.get("scannedHash"),

            log.get("result"),

            f"{log.get('reportId')}_final.pdf",
            log.get("scannedFile")

        ])


    filename = f"hospital_verification_logs_{hospital_id}.xlsx"
    filepath = os.path.join(UPLOAD_FOLDER, filename)

    wb.save(filepath)

    return send_from_directory(
        UPLOAD_FOLDER,
        filename,
        as_attachment=True
    )

# ======================================================
# MAIN HOSPITAL DASHBOARD
# ======================================================
@app.route("/main-dashboard")
@approval_required
def main_dashboard():

    hospital = get_hospital_by_id(session["hospital_id"])

    main_id = hospital["hospital_id"]

    branches = list(hospitals_collection.find({
        "parent_hospital": main_id
    }))

    reports = list(reports_collection.find({
        "hospital_id": main_id
    }))

    return render_template(
        "main_dashboard.html",
        hospital=hospital,
        branches=branches,
        reports=reports
    )

@app.route("/main/approve/<branch_id>")
def main_approve_branch(branch_id):

    if session.get("role") != "main":
        return redirect("/login")

    hospitals_collection.update_one(
        {"hospital_id": branch_id},
        {"$set": {"main_status": "approved"}}
    )

    return redirect("/main-dashboard?success=main_branch_approved")

@app.route("/main/reject/<branch_id>")
def main_reject_branch(branch_id):

    if not session.get("hospital_id") or session.get("role") != "main":
        return redirect("/login")

    hospitals_collection.update_one(
        {
            "hospital_id": branch_id,
            "parent_hospital": session["hospital_id"]
        },
        {
            "$set": {"main_status": "pending"}
        }
    )
    return redirect("/main-dashboard?success=main_branch_rejected")

@app.route("/main/delete/<branch_id>")
def main_delete_branch(branch_id):

    if not session.get("hospital_id") or session.get("role") != "main":
        return redirect("/login")

    hospitals_collection.delete_one({
        "hospital_id": branch_id,
        "parent_hospital": session["hospital_id"]
    })

    reports_collection.delete_many({
        "hospital_id": branch_id
    })

    return redirect("/main-dashboard?success=main_branch_deleted")

@app.route("/view-file/<filename>")
def view_file(filename):

    file_path = os.path.join(UPLOAD_FOLDER, filename)

    if not os.path.exists(file_path):
        return "File not found", 404

    return send_from_directory(
        UPLOAD_FOLDER,
        filename,
        as_attachment=False
    )

@app.route("/main/branch/<branch_id>/reports")
def main_branch_reports(branch_id):

    if not session.get("hospital_id") or session.get("role") != "main":
        return redirect("/login")

    # verify branch belongs to this main hospital
    branch = hospitals_collection.find_one({
        "hospital_id": branch_id,
        "parent_hospital": session["hospital_id"]
    })

    if not branch:
        return redirect("/main-dashboard")

    reports = list(reports_collection.find({
        "hospital_id": branch_id
    }).sort("created_at", -1))

    return render_template(
        "main_branch_reports.html",
        branch=branch,
        reports=reports
    )

@app.route("/main/report/delete/<report_id>")
def main_delete_report(report_id):

    if not session.get("hospital_id"):
        return redirect("/login")

    return redirect("/main-dashboard?error=delete_not_allowed")

@app.route("/main/verifications")
def main_verifications():

    if session.get("role") != "main":
        return redirect("/login")

    main_id = session["hospital_id"]

    # get all branch hospital IDs
    branches = hospitals_collection.find(
        {"parent_hospital": main_id},
        {"hospital_id": 1}
    )

    branch_ids = [b["hospital_id"] for b in branches]

    allowed_ids = [main_id] + branch_ids

    filter_type = request.args.get("type", "all")

    query = {
        "hospitalId": {"$in": allowed_ids}
    }

    if filter_type == "hospital":
        query["verifiedBy"] = {"$ne": "Public"}

    elif filter_type == "public":
        query["verifiedBy"] = "Public"

    logs = get_verification_logs(query)

    return render_template(
        "main_verifications.html",
        logs=logs,
        filter_type=filter_type,
        main_id=main_id,
        branch_ids=branch_ids
    )
@app.route("/main/branch/<branch_id>/verifications")
@approval_required
def main_branch_verifications(branch_id):

    branch = hospitals_collection.find_one({
        "hospital_id": branch_id,
        "parent_hospital": session["hospital_id"]
    })

    logs = get_verification_logs({
        "verifiedByHospitalId": branch_id
    })

    enriched_logs = []

    for log in logs:

        report_hospital = hospitals_collection.find_one({
            "hospital_id": log.get("reportHospitalId")
        })

        verifier_hospital = hospitals_collection.find_one({
            "hospital_id": log.get("verifiedByHospitalId")
        })

        report_name = ""
        report_address = ""

        if report_hospital:

            report_name = report_hospital.get("hospital_name","")
            report_address = report_hospital.get("address","")

            if report_hospital.get("branch_name"):
                report_name += f" – {report_hospital.get('branch_name')}"
            else:
                report_name += " – Main"

        verifier_name = log.get("verifiedByHospitalName","Public User")
        verifier_address = ""

        if verifier_hospital:

            verifier_address = verifier_hospital.get("address","")

            if verifier_hospital.get("branch_name"):
                verifier_name += f" – {verifier_hospital.get('branch_name')}"
            else:
                verifier_name += " – Main"

        log["reportHospitalNameFull"] = report_name
        log["reportHospitalAddress"] = report_address

        log["verifierNameFull"] = verifier_name
        log["verifierAddress"] = verifier_address

        enriched_logs.append(log)

    return render_template(
        "main_branch_verifications.html",
        branch=branch,
        logs=enriched_logs
    )
@app.route("/download/branch-verifications-excel/<branch_id>")
@approval_required
def download_branch_verifications_excel(branch_id):

    if session.get("role") != "main":
        return redirect("/login")

    logs = get_verification_logs({
        "verifiedByHospitalId": branch_id
    })

    wb = Workbook()
    ws = wb.active
    ws.title = "Branch Verification Logs"

    headers = [

        "Report ID",

        "Uploaded Hospital",
        "Hospital Address",

        "Verified By",
        "Verifier Address",

        "Verified Date",
        "Verified Time",

        "Location",
        "IP Address",

        "Original Hash",
        "Scanned Hash",

        "Result",

        "Original File",
        "Scanned File"

    ]

    ws.append(headers)

    # bold headers
    for col in range(1, len(headers)+1):
        ws.cell(row=1, column=col).font = Font(bold=True)


    for log in logs:

        report_hospital = hospitals_collection.find_one({
            "hospital_id": log.get("reportHospitalId")
        })

        verifier_hospital = hospitals_collection.find_one({
            "hospital_id": log.get("verifiedByHospitalId")
        })

        report_name = ""
        report_address = ""

        if report_hospital:

            report_name = report_hospital.get("hospital_name","")
            report_address = report_hospital.get("address","")

            if report_hospital.get("branch_name"):
                report_name += f" – {report_hospital.get('branch_name')}"
            else:
                report_name += " – Main"


        verifier_name = log.get("verifiedByHospitalName","Public User")
        verifier_address = ""

        if verifier_hospital:

            verifier_address = verifier_hospital.get("address","")

            if verifier_hospital.get("branch_name"):
                verifier_name += f" – {verifier_hospital.get('branch_name')}"
            else:
                verifier_name += " – Main"


        ws.append([

            log.get("reportId"),

            report_name,
            report_address,

            verifier_name,
            verifier_address,

            log["verifiedAt"].strftime("%d %b %Y"),
            log["verifiedAt"].strftime("%I:%M %p"),

            log.get("location"),
            log.get("ipAddress"),

            log.get("originalHash"),
            log.get("scannedHash"),

            log.get("result"),

            f"{log.get('reportId')}_final.pdf",
            log.get("scannedFile")

        ])


    filename = f"branch_verification_logs_{branch_id}.xlsx"

    filepath = os.path.join(UPLOAD_FOLDER, filename)

    wb.save(filepath)

    return send_from_directory(
        UPLOAD_FOLDER,
        filename,
        as_attachment=True
    )

@app.route("/main/my-verifications")
@approval_required
def main_my_verifications():

    logs = get_verification_logs({
        "verifiedByHospitalId": session["hospital_id"]
    })

    enriched_logs = []

    for log in logs:

        # ================= UPLOADED HOSPITAL =================
        report_hospital = hospitals_collection.find_one({
            "hospital_id": log.get("reportHospitalId")
        })

        report_name = "Unknown"
        report_address = ""

        if report_hospital:

            report_name = report_hospital.get("hospital_name", "")
            report_address = report_hospital.get("address", "")

            if report_hospital.get("branch_name"):
                report_name += f" – {report_hospital.get('branch_name')}"
            else:
                report_name += " – Main"


        # ================= VERIFIED BY =================
        verifier_hospital = hospitals_collection.find_one({
            "hospital_id": log.get("verifiedByHospitalId")
        })

        verifier_name = log.get("verifiedByHospitalName", "Public User")
        verifier_address = ""

        if verifier_hospital:

            verifier_address = verifier_hospital.get("address", "")

            if verifier_hospital.get("branch_name"):
                verifier_name += f" – {verifier_hospital.get('branch_name')}"
            else:
                verifier_name += " – Main"

        else:

            verifier_name = "Public User"
            verifier_address = "Public Network"


        # ================= ADD FIELDS =================
        log["reportHospitalNameFull"] = report_name
        log["reportHospitalAddress"] = report_address

        log["verifierNameFull"] = verifier_name
        log["verifierAddress"] = verifier_address

        enriched_logs.append(log)


    return render_template(
        "main_my_verifications.html",
        logs=enriched_logs
    )

@app.route("/download/main-my-verifications-excel")
@approval_required
def download_main_my_verifications_excel():

    hospital_id = session["hospital_id"]

    logs = get_verification_logs({
        "verifiedByHospitalId": hospital_id
    })

    wb = Workbook()
    ws = wb.active
    ws.title = "Main Hospital Verification Logs"

    headers = [

        "Report ID",

        "Uploaded Hospital",
        "Hospital Address",

        "Verified By",
        "Verifier Address",

        "Verified At",

        "Location",
        "IP Address",

        "Original Hash",
        "Scanned Hash",

        "Result",

        "Original File",
        "Scanned File"

    ]

    ws.append(headers)

    for col in range(1, len(headers)+1):
        ws.cell(row=1, column=col).font = Font(bold=True)


    for log in logs:

        # ================= UPLOADED HOSPITAL =================
        report_hospital = hospitals_collection.find_one({
            "hospital_id": log.get("reportHospitalId")
        })

        report_name = "Unknown"
        report_address = ""

        if report_hospital:

            report_name = report_hospital.get("hospital_name", "")
            report_address = report_hospital.get("address", "")

            if report_hospital.get("branch_name"):
                report_name += f" – {report_hospital.get('branch_name')}"
            else:
                report_name += " – Main"


        # ================= VERIFIED BY =================
        verifier_hospital = hospitals_collection.find_one({
            "hospital_id": log.get("verifiedByHospitalId")
        })

        verifier_name = log.get("verifiedByHospitalName", "Public User")
        verifier_address = ""

        if verifier_hospital:

            verifier_address = verifier_hospital.get("address", "")

            if verifier_hospital.get("branch_name"):
                verifier_name += f" – {verifier_hospital.get('branch_name')}"
            else:
                verifier_name += " – Main"

        else:

            verifier_name = "Public User"
            verifier_address = "Public Network"


        ws.append([

            log.get("reportId"),

            report_name,
            report_address,

            verifier_name,
            verifier_address,

            log["verifiedAt"].strftime("%d %b %Y %I:%M %p"),

            log.get("location"),
            log.get("ipAddress"),

            log.get("originalHash"),
            log.get("scannedHash"),

            log.get("result"),

            f"{log.get('reportId')}_final.pdf",
            log.get("scannedFile")

        ])


    filename = f"main_verification_logs_{hospital_id}.xlsx"
    filepath = os.path.join(UPLOAD_FOLDER, filename)

    wb.save(filepath)

    return send_from_directory(
        UPLOAD_FOLDER,
        filename,
        as_attachment=True
    )
# ======================================================
# UPLOAD REPORT
# ======================================================
@app.route("/upload", methods=["GET", "POST"])
@approval_required
def upload():

    hospital = get_hospital_by_id(session["hospital_id"])

    hospital_id = hospital["hospital_id"].upper()
    year = datetime.now().year

    # ================= FIND LAST REPORT NUMBER =================

    last_report = reports_collection.find(
        {"hospital_id": hospital_id}
    ).sort("report_id", -1).limit(1)

    last_number = 0

    for r in last_report:
        parts = r["report_id"].split("-")
        if len(parts) == 3 and parts[2].isdigit():
            last_number = int(parts[2])

    next_number = str(last_number + 1).zfill(3)

    suggested_id = f"{hospital_id}-{year}-{next_number}"

    # ================= POST =================

    if request.method == "POST":

        report_id = request.form.get("report_id").strip()
        file = request.files.get("report_file")

        # ================= VALIDATE FORMAT =================

        pattern = rf"^{hospital_id}-\d{{4}}-\d{{3}}$"

        if not re.match(pattern, report_id):

            return render_template(
                "upload.html",
                error=f"Invalid format. Use format: {hospital_id}-YYYY-001",
                suggested_id=suggested_id
            )

        # ================= PREVENT DUPLICATE =================

        existing = reports_collection.find_one({"report_id": report_id})

        if existing:

            return render_template(
                "upload.html",
                error="Report ID already exists",
                suggested_id=suggested_id
            )

        # ================= FILE CHECK =================

        if not file or file.filename == "":

            return render_template(
                "upload.html",
                error="Please select file",
                suggested_id=suggested_id
            )

        # ================= TIME =================

        ist = pytz.timezone("Asia/Kolkata")
        upload_time = datetime.now(ist)

        upload_time_str = upload_time.strftime("%d %b %Y, %I:%M %p")
        issue_date = upload_time.strftime("%Y-%m-%d")

        # ================= LOCATION =================
        ip_address = get_real_ip()

        latitude = request.form.get("latitude")
        longitude = request.form.get("longitude")

        if latitude and longitude:
            location = get_location_from_coordinates(latitude, longitude, ip_address)
        else:
            location = "Not provided"

        
        # ================= FILE SAVE =================

        ext = os.path.splitext(file.filename)[1].lower()

        original_filename = f"{report_id}_original{ext}"
        final_filename = f"{report_id}_final{ext}"

        original_path = os.path.join(UPLOAD_FOLDER, original_filename)
        final_path = os.path.join(UPLOAD_FOLDER, final_filename)

        file.save(original_path)

        # ================= QR =================

        qr_path = generate_qr(
            report_id,
            hospital["hospital_name"],
            issue_date,
            QR_DIR
        )

        if ext == ".pdf":

            attach_qr_to_pdf_safe(original_path, qr_path, final_path)
            file_type = "PDF"

        else:

            embed_qr_into_image(original_path, qr_path, final_path)
            file_type = "Image"

        # ================= HASH =================

        report_hash = generate_sha256(final_path)

        tx = store_hash(report_id, report_hash)

        if not tx:

            return render_template(
                "upload.html",
                error="Blockchain failed",
                suggested_id=suggested_id
            )

        # ================= STORE DB =================

        reports_collection.insert_one({

            "report_id": report_id,
            "hospital_id": hospital_id,
            "hospital_name": hospital["hospital_name"],
            "file_name": final_filename,
            "hash": report_hash,
            "issue_date": issue_date,
            "uploaded_at": upload_time,
            "uploaded_at_str": upload_time_str,
            "ip_address": ip_address,
            "location": location,
            "latitude": latitude,
            "longitude": longitude

        })

        qr_image_url = url_for(
            "static",
            filename=f"qrcodes/{os.path.basename(qr_path)}"
        )

        download_link = url_for(
            "download_pdf",
            filename=final_filename
        )

        return render_template(

            "upload.html",

            message="Report uploaded successfully",

            report_id=report_id,

            qr_image=qr_image_url,

            download_link=download_link,

            file_type=file_type,

            suggested_id=suggested_id,

            role=session.get("role")
        )

    # ================= GET =================

    return render_template(
        "upload.html",
        suggested_id=suggested_id
    )

@app.route("/download/<filename>")
def download_pdf(filename):
    return send_from_directory(UPLOAD_FOLDER, filename, as_attachment=True)


# ======================================================
# VERIFY REPORT
# ======================================================
def compute_pdf_sha256(path):
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(4096), b""):
            h.update(chunk)
    return h.hexdigest()

def extract_pdf_text(path):
    text = ""
    try:
        with pdfplumber.open(path) as pdf:
            for page in pdf.pages:
                text += page.extract_text() or ""
        return text
    except Exception as e:
        print("PDF text extraction error:", e)
        return ""
    
# ================= VERIFY REPORT =================
@app.route("/verify", methods=["GET", "POST"])
def verify():

    # ================= APPROVAL CHECK =================
    if session.get("hospital_id"):
        hospital = get_hospital_by_id(session["hospital_id"])
        if hospital:
            if hospital["role"] == "main":
                if hospital.get("admin_status") != "approved":
                    return render_template("pending_approval.html", hospital=hospital)
            elif hospital["role"] == "branch":
                if (
                    hospital.get("admin_status") != "approved"
                    or hospital.get("main_status") != "approved"
                ):
                    return render_template("pending_approval.html", hospital=hospital)

    # ================= GET =================
    if request.method == "GET":
        return render_template("verify.html")

    # ================= POST =================
    report_id = request.form.get("report_id")
    file = request.files.get("report_file")

    if not report_id or not file:
        return render_template("verify.html", error="Report ID and file required")

    ext = os.path.splitext(file.filename)[1].lower()

    scanned_filename = f"{report_id}_scanned_{datetime.now().strftime('%Y%m%d%H%M%S')}{ext}"
    scanned_path = os.path.join(UPLOAD_FOLDER, scanned_filename)
    file.save(scanned_path)

    uploaded_hash = compute_pdf_sha256(scanned_path)
    stored_hash = get_hash(report_id)
    meta = get_metadata(report_id)

    formatted_changes = []

    # ================= RESULT =================
    if stored_hash and uploaded_hash == stored_hash:

        result = "Genuine"
        message = "Report is genuine"
        status = "success"

    else:

        result = "Fake"
        message = "Report is fake or modified"
        status = "error"

        # ================= FIND DIFFERENCES =================
        original_report = reports_collection.find_one({"report_id": report_id})

        if original_report:

            original_path = os.path.join(
                UPLOAD_FOLDER,
                original_report.get("file_name")
            )

            original_text = extract_pdf_text(original_path)
            uploaded_text = extract_pdf_text(scanned_path)

            diff = list(difflib.ndiff(
                original_text.splitlines(),
                uploaded_text.splitlines()
            ))

            removed_lines = []
            added_lines = []

            for line in diff:
                if line.startswith("- "):
                    removed_lines.append(line[2:])
                elif line.startswith("+ "):
                    added_lines.append(line[2:])

            # Pair removed and added lines
            for i in range(min(len(removed_lines), len(added_lines))):
                formatted_changes.append({
                    "original": removed_lines[i],
                    "modified": added_lines[i]
                })

    # ================= LOCATION =================
    latitude = request.form.get("latitude")
    longitude = request.form.get("longitude")
    ip_address = get_real_ip()

    if latitude and longitude:
        location = get_location_from_coordinates(latitude, longitude, ip_address)
    else:
        location = get_location_from_ip(ip_address)

    # ================= VERIFIED BY =================
    if session.get("admin"):
        verified_by = "Admin"
        verifier_id = "ADMIN"
        verifier_name = "System Admin"

    elif session.get("role") == "main":
        verified_by = "Main Hospital"
        verifier_id = session["hospital_id"]
        verifier_name = session["hospital_name"]

    elif session.get("role") == "branch":
        verified_by = "Hospital"
        verifier_id = session["hospital_id"]
        verifier_name = session["hospital_name"]

    else:
        verified_by = "Public"
        verifier_id = None
        verifier_name = "Public User"

    # ================= SAVE LOG =================
    log_verification(
        report_id=report_id,
        report_hospital_id=meta["hospital_id"] if meta else None,
        report_hospital_name=meta.get("hospital_name") if meta else None,
        verified_by=verified_by,
        verified_by_hospital_id=verifier_id,
        verified_by_hospital_name=verifier_name,
        branch_name=session.get("branch_name"),
        original_hash=stored_hash,
        scanned_hash=uploaded_hash,
        scanned_file=scanned_filename,
        result=result,
        ip_address=ip_address,
        location=location,
        latitude=latitude,
        longitude=longitude
    )

    # ================= REDIRECT LOGIC =================
    if session.get("admin"):
        redirect_url = url_for("admin_dashboard")

    elif session.get("role") == "main":
        redirect_url = url_for("main_dashboard")

    elif session.get("role") == "branch":
        redirect_url = url_for("hospital_dashboard")

    else:
        redirect_url = url_for("home")

    return render_template(
        "verify.html",
        message=message,
        status=status,
        changed_content=formatted_changes,
        redirect_url=redirect_url
    )
    
@app.route("/view-original/<report_id>")
def view_original(report_id):

    if not session.get("hospital_id") and not session.get("admin"):
        return redirect("/login")

    report = reports_collection.find_one({"report_id": report_id})

    hospital = hospitals_collection.find_one({
        "hospital_id": report.get("hospital_id")
    })

    issued_display = hospital.get("hospital_name", "")

    if hospital.get("branch_name"):
        issued_display += f" – {hospital.get('branch_name')}"
    else:
        issued_display += " – Main"

    report["issued_display"] = issued_display
    report["view_type"] = "Original Report"

    return render_template("view_report.html", report=report)

@app.route("/view-scanned/<filename>")
def view_scanned(filename):

    if not session.get("hospital_id") and not session.get("admin"):
        return redirect("/login")

    report_id = filename.split("_scanned")[0]

    original = reports_collection.find_one({"report_id": report_id})

    hospital = hospitals_collection.find_one({
        "hospital_id": original.get("hospital_id")
    })

    issued_display = hospital.get("hospital_name", "")

    if hospital.get("branch_name"):
        issued_display += f" – {hospital.get('branch_name')}"
    else:
        issued_display += " – Main"

    report = {
        "report_id": report_id,
        "issued_display": issued_display,
        "issue_date": original.get("issue_date"),
        "file_name": filename,
        "view_type": "Scanned Report"
    }

    return render_template("view_report.html", report=report)

@app.route("/qr-status-by-token/<token>")
def qr_status_by_token(token):

    try:

        # STEP 1: decode token
        report_id = get_report_id_from_token(token)

        if not report_id:
            return {"status": "invalid"}

        # STEP 2: get metadata
        meta = get_metadata(report_id)

        if not meta:
            return {"status": "not_found"}

        # STEP 3: check blockchain hash
        hash_val = get_hash(report_id)

        if not hash_val:
            return {"status": "not_found"}

        # STEP 4: get hospital info
        hospital = hospitals_collection.find_one({
            "hospital_id": meta.get("hospital_id")
        })

        branch_name = ""
        main_hospital = ""

        if hospital:

            branch_name = hospital.get("branch_name", "")

            parent_id = hospital.get("parent_hospital")

            if parent_id:

                main = hospitals_collection.find_one({
                    "hospital_id": parent_id
                })

                if main:
                    main_hospital = main.get("hospital_name", "")

        # STEP 5: return success
        return {

            "status": "registered",

            "report_id": report_id,

            "hospital": meta.get("hospital_name", ""),

            "branch": branch_name,

            "main_hospital": main_hospital,

            "issue_date": meta.get("issue_date", "")

        }

    except Exception as e:

        print("QR STATUS ERROR:", str(e))

        return {"status": "error"}

@app.route("/admin/main/<main_id>/verifications")
def admin_main_hospital_verifications(main_id):

    if not session.get("admin"):
        return redirect("/admin/login")

    hospital = hospitals_collection.find_one({
        "hospital_id": main_id,
        "role": "main"
    })

    if not hospital:
        return "Hospital not found"

    logs = get_verification_logs({
        "verifiedByHospitalId": main_id
    })

    enriched_logs = []

    for log in logs:

        # ================= UPLOADED HOSPITAL =================
        report_hospital = hospitals_collection.find_one({
            "hospital_id": log.get("reportHospitalId")
        })

        report_name = "Unknown"
        report_address = ""

        if report_hospital:

            report_name = report_hospital.get("hospital_name", "")
            report_address = report_hospital.get("address", "")

            if report_hospital.get("branch_name"):
                report_name += f" – {report_hospital.get('branch_name')}"
            else:
                report_name += " – Main"


        # ================= VERIFIED BY =================
        verifier_hospital = hospitals_collection.find_one({
            "hospital_id": log.get("verifiedByHospitalId")
        })

        verifier_name = log.get("verifiedByHospitalName", "Public User")
        verifier_address = ""

        if log.get("verifiedBy") == "Admin":

            verifier_name = "System Admin"
            verifier_address = "System"

        elif verifier_hospital:

            verifier_address = verifier_hospital.get("address", "")

            if verifier_hospital.get("branch_name"):
                verifier_name += f" – {verifier_hospital.get('branch_name')}"
            else:
                verifier_name += " – Main"

        else:

            verifier_name = "Public User"
            verifier_address = "Public Network"


        # ADD SAME FIELDS AS MY VERIFICATION LOGS
        log["reportHospitalNameFull"] = report_name
        log["reportHospitalAddress"] = report_address

        log["verifiedByHospitalName"] = verifier_name
        log["verifierAddress"] = verifier_address

        enriched_logs.append(log)

    return render_template(
        "admin/admin_main_verifications.html",
        hospital=hospital,
        logs=enriched_logs
    )

@app.route("/download/admin-main-verifications-excel/<main_id>")
def download_admin_main_verifications_excel(main_id):

    if not session.get("admin"):
        return redirect("/admin/login")

    logs = get_verification_logs({
        "verifiedByHospitalId": main_id
    })

    wb = Workbook()
    ws = wb.active
    ws.title = "Main Verification Logs"

    headers = [
        "Report ID",
        "Uploaded Hospital",
        "Hospital Address",
        "Verified By",
        "Verifier Address",
        "Verified At",
        "Location",
        "IP Address",
        "Original Hash",
        "Scanned Hash",
        "Result"
    ]

    ws.append(headers)

    for col in range(1, len(headers)+1):
        ws.cell(row=1, column=col).font = Font(bold=True)

    for log in logs:

        # uploaded hospital
        report_hospital = hospitals_collection.find_one({
            "hospital_id": log.get("reportHospitalId")
        })

        report_name = "Unknown"
        report_address = ""

        if report_hospital:

            report_name = report_hospital.get("hospital_name", "")
            report_address = report_hospital.get("address", "")

            if report_hospital.get("branch_name"):
                report_name += f" – {report_hospital.get('branch_name')}"
            else:
                report_name += " – Main"

        # verifier
        verifier_name = log.get("verifiedByHospitalName", "Public User")
        verifier_address = ""

        verifier_hospital = hospitals_collection.find_one({
            "hospital_id": log.get("verifiedByHospitalId")
        })

        if log.get("verifiedBy") == "Admin":

            verifier_name = "System Admin"
            verifier_address = "System"

        elif verifier_hospital:

            verifier_address = verifier_hospital.get("address", "")

            if verifier_hospital.get("branch_name"):
                verifier_name += f" – {verifier_hospital.get('branch_name')}"
            else:
                verifier_name += " – Main"

        else:

            verifier_name = "Public User"
            verifier_address = "Public Network"

        ws.append([
            log.get("reportId"),
            report_name,
            report_address,
            verifier_name,
            verifier_address,
            log["verifiedAt"].strftime("%d %b %Y %I:%M %p"),
            log.get("location"),
            log.get("ipAddress"),
            log.get("originalHash"),
            log.get("scannedHash"),
            log.get("result")
        ])

    filename = f"admin_main_logs_{main_id}.xlsx"
    path = os.path.join(UPLOAD_FOLDER, filename)

    wb.save(path)

    return send_from_directory(
        UPLOAD_FOLDER,
        filename,
        as_attachment=True
    )

@app.route("/admin/branch/<branch_id>/verifications")
def admin_branch_verifications(branch_id):

    if not session.get("admin"):
        return redirect("/admin/login")

    branch = hospitals_collection.find_one({
        "hospital_id": branch_id,
        "role": "branch"
    })

    if not branch:
        return "Branch not found"

    logs = get_verification_logs({
        "verifiedByHospitalId": branch_id
    })

    enriched_logs = []

    for log in logs:

        # UPLOADED HOSPITAL
        report_hospital = hospitals_collection.find_one({
            "hospital_id": log.get("reportHospitalId")
        })

        report_name = "Unknown"
        report_address = ""

        if report_hospital:

            report_name = report_hospital.get("hospital_name", "")
            report_address = report_hospital.get("address", "")

            if report_hospital.get("branch_name"):
                report_name += f" – {report_hospital.get('branch_name')}"
            else:
                report_name += " – Main"


        # VERIFIED BY HOSPITAL
        verifier_hospital = hospitals_collection.find_one({
            "hospital_id": log.get("verifiedByHospitalId")
        })

        verifier_name = log.get("verifiedByHospitalName", "Public User")
        verifier_address = ""

        if log.get("verifiedBy") == "Admin":

            verifier_name = "System Admin"
            verifier_address = "System"

        elif verifier_hospital:

            verifier_address = verifier_hospital.get("address", "")

            if verifier_hospital.get("branch_name"):
                verifier_name += f" – {verifier_hospital.get('branch_name')}"
            else:
                verifier_name += " – Main"

        else:

            verifier_name = "Public User"
            verifier_address = "Public Network"


        # ADD EXTRA FIELDS (your template still uses log.reportId etc)
        log["reportHospitalNameFull"] = report_name
        log["reportHospitalAddress"] = report_address

        log["verifiedByHospitalName"] = verifier_name
        log["verifierAddress"] = verifier_address

        enriched_logs.append(log)


    return render_template(
        "admin/admin_branch_verifications.html",
        branch=branch,
        logs=enriched_logs
    )
@app.route("/download/admin-branch-verifications-excel/<branch_id>")
def download_admin_branch_verifications_excel(branch_id):

    if not session.get("admin"):
        return redirect("/admin/login")

    logs = get_verification_logs({
        "verifiedByHospitalId": branch_id
    })

    wb = Workbook()
    ws = wb.active
    ws.title = "Branch Verification Logs"

    headers = [

        "Report ID",

        "Uploaded Hospital",
        "Hospital Address",

        "Verified By",
        "Verifier Address",

        "Verified Date",
        "Verified Time",

        "Location",
        "IP Address",

        "Original Hash",
        "Scanned Hash",

        "Result",

        "Original File",
        "Scanned File"
    ]

    ws.append(headers)

    for col in range(1, len(headers)+1):
        ws.cell(row=1, column=col).font = Font(bold=True)


    for log in logs:

        report_hospital = hospitals_collection.find_one({
            "hospital_id": log.get("reportHospitalId")
        })

        verifier_hospital = hospitals_collection.find_one({
            "hospital_id": log.get("verifiedByHospitalId")
        })

        report_name = ""
        report_address = ""

        if report_hospital:

            report_name = report_hospital.get("hospital_name", "")
            report_address = report_hospital.get("address", "")

            if report_hospital.get("branch_name"):
                report_name += f" – {report_hospital.get('branch_name')}"
            else:
                report_name += " – Main"


        verifier_name = log.get("verifiedByHospitalName", "Public User")
        verifier_address = ""

        if verifier_hospital:

            verifier_address = verifier_hospital.get("address", "")

            if verifier_hospital.get("branch_name"):
                verifier_name += f" – {verifier_hospital.get('branch_name')}"
            else:
                verifier_name += " – Main"

        else:

            verifier_name = "Public User"
            verifier_address = "Public Network"


        ws.append([

            log.get("reportId"),

            report_name,
            report_address,

            verifier_name,
            verifier_address,

            log["verifiedAt"].strftime("%d %b %Y"),
            log["verifiedAt"].strftime("%I:%M %p"),

            log.get("location"),
            log.get("ipAddress"),

            log.get("originalHash"),
            log.get("scannedHash"),

            log.get("result"),

            f"{log.get('reportId')}_final.pdf",
            log.get("scannedFile")

        ])


    filename = f"admin_branch_logs_{branch_id}.xlsx"

    filepath = os.path.join(UPLOAD_FOLDER, filename)

    wb.save(filepath)

    return send_from_directory(
        UPLOAD_FOLDER,
        filename,
        as_attachment=True
    )

@app.route("/hospital/report/<report_id>/verifications")
def report_verification_logs(report_id):

    if not session.get("hospital_id") and not session.get("admin"):
        return redirect("/login")

    filter_type = request.args.get("type", "all")

    query = {"reportId": report_id}

    if filter_type == "hospital":
        query["verifiedBy"] = "Hospital"

    elif filter_type == "public":
        query["verifiedBy"] = "Public"

    elif filter_type == "admin":
        query["verifiedBy"] = {"$in": ["Admin", "Main Hospital"]}

    logs = get_verification_logs(query)

    enriched_logs = []

    for log in logs:

        # ================= UPLOADED HOSPITAL =================
        report_hospital = hospitals_collection.find_one({
            "hospital_id": log.get("reportHospitalId")
        })

        report_name = "Unknown"
        report_address = ""

        if report_hospital:

            report_name = report_hospital.get("hospital_name", "")
            report_address = report_hospital.get("address", "")

            if report_hospital.get("branch_name"):
                report_name += f" – {report_hospital.get('branch_name')}"
            else:
                report_name += " – Main"


        # ================= VERIFIED BY =================
        verifier_hospital = hospitals_collection.find_one({
            "hospital_id": log.get("verifiedByHospitalId")
        })

        verifier_name = log.get("verifiedByHospitalName", "Public User")
        verifier_address = ""

        if log.get("verifiedBy") == "Admin":

            verifier_name = "System Admin"
            verifier_address = "System"

        elif verifier_hospital:

            verifier_address = verifier_hospital.get("address", "")

            if verifier_hospital.get("branch_name"):
                verifier_name += f" – {verifier_hospital.get('branch_name')}"
            else:
                verifier_name += " – Main"

        else:

            verifier_name = "Public User"
            verifier_address = "Public Network"


        # ================= ADD FIELDS =================
        log["reportHospitalNameFull"] = report_name
        log["reportHospitalAddress"] = report_address

        log["verifierNameFull"] = verifier_name
        log["verifierAddress"] = verifier_address

        enriched_logs.append(log)

    return render_template(
        "report_verifications.html",
        report_id=report_id,
        logs=enriched_logs,
        filter_type=filter_type
    )

@app.route("/view-report/<report_id>")
def view_report(report_id):

    report = reports_collection.find_one({"report_id": report_id})

    if not report:
        return "Report not found"

    hospital = hospitals_collection.find_one({
        "hospital_id": report.get("hospital_id")
    })

    issued_display = "Unknown Hospital"

    if hospital:
        issued_display = hospital.get("hospital_name", "")
        if hospital.get("branch_name"):
            issued_display += f" – {hospital.get('branch_name')}"
        else:
            issued_display += " – Main"

    report["issued_display"] = issued_display
    report["view_type"] = "Medical Report"

    return render_template("view_report.html", report=report)

@app.route("/download/report-logs-excel/<report_id>")
def download_report_logs_excel(report_id):

    logs = get_verification_logs({"reportId": report_id})

    wb = Workbook()
    ws = wb.active
    ws.title = "Verification Logs"

    headers = [

        "Report ID",

        "Uploaded Hospital",
        "Hospital Address",

        "Verified By",
        "Verifier Address",

        "Verified At",

        "Location",
        "IP Address",

        "Original Hash",
        "Scanned Hash",

        "Result",

        "Original File",
        "Scanned File"
    ]

    ws.append(headers)

    for col in range(1, len(headers)+1):
        ws.cell(row=1, column=col).font = Font(bold=True)


    for log in logs:

        # ================= UPLOADED HOSPITAL =================
        report_hospital = hospitals_collection.find_one({
            "hospital_id": log.get("reportHospitalId")
        })

        report_name = "Unknown"
        report_address = ""

        if report_hospital:

            report_name = report_hospital.get("hospital_name", "")
            report_address = report_hospital.get("address", "")

            if report_hospital.get("branch_name"):
                report_name += f" – {report_hospital.get('branch_name')}"
            else:
                report_name += " – Main"


        # ================= VERIFIED BY =================
        verifier_hospital = hospitals_collection.find_one({
            "hospital_id": log.get("verifiedByHospitalId")
        })

        verifier_name = log.get("verifiedByHospitalName", "Public User")
        verifier_address = ""

        if log.get("verifiedBy") == "Admin":

            verifier_name = "System Admin"
            verifier_address = "System"

        elif verifier_hospital:

            verifier_address = verifier_hospital.get("address", "")

            if verifier_hospital.get("branch_name"):
                verifier_name += f" – {verifier_hospital.get('branch_name')}"
            else:
                verifier_name += " – Main"

        else:

            verifier_name = "Public User"
            verifier_address = "Public Network"


        ws.append([

            log.get("reportId"),

            report_name,
            report_address,

            verifier_name,
            verifier_address,

            log["verifiedAt"].strftime("%d %b %Y %I:%M %p"),

            log.get("location"),
            log.get("ipAddress"),

            log.get("originalHash"),
            log.get("scannedHash"),

            log.get("result"),

            f"{log.get('reportId')}_final.pdf",
            log.get("scannedFile")

        ])


    filename = f"verification_logs_{report_id}.xlsx"
    filepath = os.path.join(UPLOAD_FOLDER, filename)

    wb.save(filepath)

    return send_from_directory(
        UPLOAD_FOLDER,
        filename,
        as_attachment=True
    )

# ======================================================
# APPROVAL CHECK FUNCTION
# ======================================================
def is_approved(hospital):

    if hospital["role"] == "main":
        return hospital.get("admin_status") == "approved"

    if hospital["role"] == "branch":
        return (
            hospital.get("admin_status") == "approved"
            and hospital.get("main_status") == "approved"
        )

    return False


# ======================================================
# MANUAL VERIFY PAGE (QR SCANNER PAGE)
# ======================================================
@app.route("/manual-verify")
def manual_verify_page():
    return render_template("manual_verify.html")


# ======================================================
# MANUAL VERIFY BACKEND (QR TOKEN CHECK)
# ======================================================
@app.route("/manual-verify/<token>")
def manual_verify_token(token):

    report_id = get_report_id_from_token(token)

    meta = get_metadata(report_id) if report_id else None
    stored_hash = get_hash(report_id) if report_id else None

    report = reports_collection.find_one(
        {"report_id": report_id}
    ) if report_id else None

    # ================= RESULT LOGIC =================
    if meta and stored_hash and report:

        log_result = "Genuine"      # Saved in DB
        ui_status = "matched"       # Sent to frontend
        file_name = report.get("file_name")

    else:

        log_result = "Fake"
        ui_status = "not_matched"
        file_name = None


    # ================= LOCATION =================
    latitude = request.args.get("lat")
    longitude = request.args.get("lon")

    ip_address = get_real_ip()

    if latitude and longitude:
        location = get_location_from_coordinates(latitude, longitude, ip_address)
    else:
        location = get_location_from_ip(ip_address)


    # ================= VERIFIED BY =================
    if session.get("admin"):

        verified_by = "Admin"
        verifier_id = "ADMIN"
        verifier_name = "System Admin"

    elif session.get("hospital_id"):

        if session.get("role") == "main":
            verified_by = "Main Hospital"
        else:
            verified_by = "Hospital"

        verifier_id = session["hospital_id"]
        verifier_name = session["hospital_name"]

    else:

        verified_by = "Public"
        verifier_id = None
        verifier_name = "Public User"


    # ================= SAVE LOG =================
    log_verification(

        report_id=report_id,

        report_hospital_id=meta.get("hospital_id") if meta else None,
        report_hospital_name=meta.get("hospital_name") if meta else None,

        verified_by=verified_by,
        verified_by_hospital_id=verifier_id,
        verified_by_hospital_name=verifier_name,

        branch_name=session.get("branch_name"),

        original_hash=stored_hash,
        scanned_hash="Manual Verification",
        scanned_file="Manual Verification",

        result=log_result,   # 🔥 IMPORTANT

        ip_address=ip_address,
        location=location,
        latitude=latitude,
        longitude=longitude
    )

    return {
        "status": ui_status,
        "report_id": report_id
    }

# ======================================================
# RUN (LOCAL + DEPLOYMENT SUPPORT)
# ======================================================
if __name__ == "__main__":

    # Detect environment
    port = int(os.environ.get("PORT", 5000))

    # Run locally or on deployment
    app.run(
        host="0.0.0.0",   # Required for deployment
        port=port,        # Deployment platforms assign PORT
        debug=True        # Change to False in production if needed
    )